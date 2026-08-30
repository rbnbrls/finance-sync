"""Read-only importer for official DEGIRO pension-account exports.

The connector deliberately contains no network client or authentication flow.
It accepts transaction, account-statement and portfolio exports in CSV, XLSX
or XLS format and detects the report type from its contents.
"""

from __future__ import annotations

import csv
import hashlib
import io
import os
import re
import unicodedata
from collections import Counter
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from finance_sync.connectors.base import Connector
from finance_sync.connectors.exceptions import PermanentError
from finance_sync.connectors.models import (
    ConnectorConfig,
    RawAccount,
    RawHolding,
    RawTransaction,
    SecurityReference,
)

_AMSTERDAM = ZoneInfo("Europe/Amsterdam")
_SUPPORTED = {".csv", ".xlsx", ".xls"}
_PDF_MESSAGE = (
    "PDF-bestanden worden niet ondersteund. Exporteer in DEGIRO het "
    "transactieoverzicht, rekeningoverzicht en portefeuilleoverzicht als "
    "CSV of Excel en probeer het opnieuw."
)


@dataclass(slots=True)
class ImportValidationReport:
    """Validation details retained after every attempted import."""

    files: int = 0
    rows_read: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    report_types: set[str] = field(default_factory=lambda: set[str]())
    errors: list[str] = field(default_factory=lambda: list[str]())
    warnings: list[str] = field(default_factory=lambda: list[str]())

    @property
    def successful(self) -> bool:
        return not self.errors


@dataclass(slots=True)
class _Report:
    kind: str
    headers: list[str]
    rows: list[list[Any]]
    source: Path


def _key(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(
        r"[^a-z0-9]+", "", text.encode("ascii", "ignore").decode().lower()
    )


def _clean(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value).replace("\xa0", " ").strip()


def _decimal(value: object, *, required: bool = False) -> Decimal | None:
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    text = _clean(value)
    if not text or text in {"-", "--"}:
        if required:
            message = "bedrag of aantal ontbreekt"
            raise ValueError(message)
        return None
    text = re.sub(r"[€$£\s]", "", text)
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")
    try:
        result = Decimal(text)
    except InvalidOperation as exc:
        message = f"ongeldig getal {value!r}"
        raise ValueError(message) from exc
    return -result if negative else result


def _currency(value: object, default: str = "EUR") -> str:
    code = _clean(value).upper()
    return code if re.fullmatch(r"[A-Z]{3}", code) else default


def _parse_datetime(date_value: object, time_value: object = "") -> datetime:
    if isinstance(date_value, datetime):
        parsed = date_value
    elif isinstance(date_value, date):
        parsed = datetime.combine(date_value, datetime.min.time())
    else:
        raw_date = _clean(date_value)
        raw_time = _clean(time_value) or "00:00"
        parsed = None
        for fmt in (
            "%d-%m-%Y %H:%M:%S",
            "%d-%m-%Y %H:%M",
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y %H:%M",
        ):
            try:
                parsed = datetime.strptime(f"{raw_date} {raw_time}", fmt)
                break
            except ValueError:
                continue
        if parsed is None:
            message = f"ongeldige datum/tijd {raw_date!r} {raw_time!r}"
            raise ValueError(message)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=_AMSTERDAM)
    return parsed.astimezone(UTC)


def _hash(*values: object) -> str:
    normalized = "\x1f".join(_clean(v).casefold() for v in values)
    return hashlib.sha256(normalized.encode()).hexdigest()


class _Row:
    """Duplicate-header-safe accessor for DEGIRO's repeated Currency fields."""

    def __init__(self, headers: list[str], values: list[Any]) -> None:
        self.headers = headers
        self.values = values + [""] * max(0, len(headers) - len(values))
        self.keys = [_key(header) for header in headers]

    def get(self, *aliases: str, occurrence: int = 0) -> Any:
        wanted = {_key(alias) for alias in aliases}
        indexes = [i for i, key in enumerate(self.keys) if key in wanted]
        if occurrence < len(indexes):
            return self.values[indexes[occurrence]]
        return ""

    def at(self, index: int) -> Any:
        return self.values[index] if index < len(self.values) else ""

    def after(self, *aliases: str, offset: int = 1) -> Any:
        """Return a cell after a named column, including unlabeled columns."""
        wanted = {_key(alias) for alias in aliases}
        for index, key in enumerate(self.keys):
            if key in wanted:
                return self.at(index + offset)
        return ""


class DegiroPensionConnector(Connector):
    """Import official DEGIRO exports for a Dutch pension account."""

    display_name = "DEGIRO Pensioen"
    sdk_version = "0.1.0"
    supported_resources = frozenset({"accounts", "transactions", "holdings"})
    ingestion_methods = ("file",)
    import_wizard = {
        "files": [
            {"key": "account", "label": "Accountoverzicht", "required": True},
            {"key": "transactions", "label": "Transacties", "required": True},
            {"key": "portfolio", "label": "Portefeuille", "required": True},
        ],
        "accept": [".csv", ".xlsx", ".xls"],
        "preview": True,
    }
    rate_limit_policy = None

    def __init__(self, config: ConnectorConfig) -> None:
        super().__init__(config)
        self.validation_report = ImportValidationReport()
        self._reports: list[_Report] | None = None
        self._transactions: list[RawTransaction] | None = None
        self._holdings: list[RawHolding] | None = None
        self._snapshot_at: datetime | None = None
        self._portfolio_total: Decimal | None = None
        self._cash_total = Decimal(0)
        self._trade_order_ids: set[str] = set()

    @property
    def name(self) -> str:
        return "degiro_pension"

    @property
    def external_account_id(self) -> str:
        account_key = _clean(self.config.options.get("account_key"))
        if not account_key:
            account_key = "|".join(str(p.resolve()) for p in self._paths())
        return f"degiro-pension-{_hash('account', account_key)[:24]}"

    @property
    def report_summaries(self) -> list[dict[str, object]]:
        """Return safe, path-free metadata for upload previews."""
        self._ensure_loaded()
        return [
            {
                "report_type": report.kind,
                "file_name": report.source.name,
                "rows": len(report.rows),
                "columns": len(report.headers),
            }
            for report in self._reports or []
        ]

    async def authenticate(self) -> None:
        paths = self._paths()
        if not paths:
            message = (
                "Selecteer de drie officiële DEGIRO-exports (CSV of Excel) "
                "of geef een map met exports op."
            )
            raise PermanentError(message)
        for path in paths:
            if path.suffix.lower() == ".pdf":
                raise PermanentError(_PDF_MESSAGE)
            if path.suffix.lower() not in _SUPPORTED:
                message = (
                    f"{path.name}: dit bestandstype wordt niet ondersteund; "
                    "gebruik CSV of Excel."
                )
                raise PermanentError(message)
            if not path.is_file() or not os.access(path, os.R_OK):
                message = f"Kan exportbestand {path} niet lezen."
                raise PermanentError(message)
        self._load_reports()
        self._authenticated = True

    async def fetch_accounts(self) -> list[RawAccount]:
        self._ensure_loaded()
        balance = self._portfolio_total
        return [
            RawAccount(
                external_account_id=self.external_account_id,
                name=_clean(
                    self.config.options.get("account_name", "DEGIRO Pensioen")
                ),
                account_type="investment",
                account_subtype="nl_lijfrente",
                currency_code="EUR",
                current_balance=balance,
                available_balance=self._cash_total,
                iso_currency_code="EUR",
                provider_metadata={
                    "source": "official_user_export",
                    "snapshot_at": self._snapshot_at.isoformat()
                    if self._snapshot_at
                    else None,
                    "cash_included_in_current_balance": True,
                    "supports_multi_currency_cash": False,
                },
            )
        ]

    async def fetch_transactions(
        self,
        since: datetime,
        *,
        account_id: str | None = None,
        limit: int | None = None,
    ) -> list[RawTransaction]:
        self._ensure_loaded()
        if account_id is not None and account_id != self.external_account_id:
            return []
        since_utc = (
            since.astimezone(UTC) if since.tzinfo else since.replace(tzinfo=UTC)
        )
        result = [
            t for t in self._transactions or [] if t.occurred_at >= since_utc
        ]
        return result[:limit] if limit is not None else result

    async def fetch_holdings(
        self, *, account_id: str | None = None
    ) -> list[RawHolding]:
        self._ensure_loaded()
        if account_id is not None and account_id != self.external_account_id:
            return []
        return list(self._holdings or [])

    def _ensure_loaded(self) -> None:
        if self._reports is None:
            self._load_reports()

    def _paths(self) -> list[Path]:
        paths: list[Path] = []
        raw_paths = self.config.options.get("export_paths", [])
        if isinstance(raw_paths, str):
            raw_paths = [raw_paths]
        paths.extend(Path(str(raw)).expanduser() for raw in raw_paths)
        single = self.config.options.get("export_path")
        if single:
            paths.append(Path(str(single)).expanduser())
        directory = self.config.options.get("export_directory")
        if directory:
            base = Path(str(directory)).expanduser()
            if not base.is_dir():
                return [base]
            paths.extend(
                path
                for path in sorted(base.iterdir())
                if path.is_file()
                and path.suffix.lower() in _SUPPORTED | {".pdf"}
            )
        return list(dict.fromkeys(paths))

    def _load_reports(self) -> None:
        report = ImportValidationReport()
        reports: list[_Report] = []
        self._snapshot_at = None
        self._portfolio_total = None
        self._cash_total = Decimal(0)
        self._trade_order_ids = set()
        try:
            for path in self._paths():
                headers, rows = self._read_table(path)
                kind = self._detect_report(headers, rows)
                reports.append(_Report(kind, headers, rows, path))
                report.files += 1
                report.rows_read += len(rows)
                report.report_types.add(kind)
            transactions: list[RawTransaction] = []
            holdings: list[RawHolding] = []
            for source in reports:
                if source.kind == "transactions":
                    for values in source.rows:
                        order_id = _clean(
                            _Row(source.headers, values).get(
                                "Order ID", "Order Id", "OrderID"
                            )
                        )
                        if order_id:
                            self._trade_order_ids.add(order_id)
            for source in reports:
                if source.kind == "transactions":
                    transactions.extend(
                        self._parse_transactions(source, report)
                    )
                elif source.kind == "account_statement":
                    transactions.extend(self._parse_statement(source, report))
                else:
                    holdings.extend(self._parse_portfolio(source, report))
            if report.errors:
                preview = "; ".join(report.errors[:5])
                message = (
                    f"DEGIRO-import bevat {len(report.errors)} ongeldige "
                    f"regel(s): {preview}"
                )
                raise PermanentError(message)
            self._transactions = self._deduplicate(transactions)
            self._holdings = holdings
            report.rows_imported = len(self._transactions) + len(holdings)
            self._reports = reports
        except PermanentError:
            self.validation_report = report
            raise
        except Exception as exc:
            report.errors.append(str(exc))
            self.validation_report = report
            message = f"DEGIRO-export kon niet worden gelezen: {exc}"
            raise PermanentError(message) from exc
        self.validation_report = report

    def _read_table(self, path: Path) -> tuple[list[str], list[list[Any]]]:
        suffix = path.suffix.lower()
        if suffix == ".csv":
            data = path.read_bytes()
            text = None
            for encoding in ("utf-8-sig", "cp1252", "latin-1"):
                try:
                    text = data.decode(encoding)
                    break
                except UnicodeDecodeError:
                    continue
            if text is None:
                message = f"{path.name}: onbekende tekstcodering"
                raise ValueError(message)
            sample = text[:8192]
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
                delimiter = dialect.delimiter
            except csv.Error:
                delimiter = (
                    ";" if sample.count(";") > sample.count(",") else ","
                )
            table: list[list[Any]] = [
                list(row)
                for row in csv.reader(io.StringIO(text), delimiter=delimiter)
            ]
        elif suffix == ".xlsx":
            from openpyxl import load_workbook

            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                sheet = workbook.active
                if sheet is None:
                    message = f"{path.name}: het werkboek heeft geen werkblad"
                    raise ValueError(message)
                # DEGIRO's XLSX exports currently declare ``dimension A1``
                # even when the sheet contains many populated cells. In
                # read-only mode openpyxl trusts that stale dimension and
                # would return only the first cell. Force a scan of the real
                # worksheet bounds before reading the table.
                if sheet.calculate_dimension() in {"A1", "A1:A1"}:
                    sheet.reset_dimensions()
                table = [list(row) for row in sheet.iter_rows(values_only=True)]
            finally:
                workbook.close()
        else:
            import xlrd

            book = xlrd.open_workbook(str(path), on_demand=True)
            sheet = book.sheet_by_index(0)
            table = [
                list(sheet.row_values(index)) for index in range(sheet.nrows)
            ]
            book.release_resources()
        table = [row for row in table if any(_clean(cell) for cell in row)]
        if not table:
            message = f"{path.name}: het bestand is leeg"
            raise ValueError(message)
        header_index = self._header_index(table)
        headers = [_clean(value) for value in table[header_index]]
        rows = table[header_index + 1 :]
        return headers, rows

    @staticmethod
    def _header_index(table: list[list[Any]]) -> int:
        markers = {
            "datum",
            "date",
            "product",
            "isin",
            "omschrijving",
            "description",
        }
        best_index, best_score = 0, -1
        for index, row in enumerate(table[:25]):
            keys = {_key(value) for value in row}
            score = len(keys & markers) * 10 + sum(bool(key) for key in keys)
            if score > best_score:
                best_index, best_score = index, score
        return best_index

    @staticmethod
    def _detect_report(headers: list[str], rows: list[list[Any]]) -> str:
        keys = {_key(header) for header in headers}
        transaction_markers = {
            "orderid",
            "orderidd",
            "venue",
            "uitvoeringsplaats",
            "quantity",
            "aantal",
            "price",
            "koers",
            "transactionandorthirdpartyfees",
            "transactiekosten",
        }
        statement_markers = {
            "omschrijving",
            "description",
            "valuedate",
            "valutadatum",
            "balance",
            "saldo",
            "mutatie",
            "change",
        }
        portfolio_markers = {
            "lokalewaarde",
            "localvalue",
            "waardeineur",
            "valueineur",
            "slotkoers",
            "closingprice",
            "gak",
            "averageprice",
        }
        scores = {
            "transactions": len(keys & transaction_markers),
            "account_statement": len(keys & statement_markers),
            "portfolio": len(keys & portfolio_markers),
        }
        kind, score = max(scores.items(), key=lambda item: item[1])
        if score < 2:
            columns = len(headers)
            if columns in {12, 14, 18} and {"product", "isin"} <= keys:
                kind = (
                    "transactions" if {"datum", "date"} & keys else "portfolio"
                )
            else:
                message = (
                    "rapporttype niet herkend; gebruik een ongewijzigde "
                    "officiële DEGIRO-export"
                )
                raise ValueError(message)
        if not rows:
            return kind
        return kind

    def _parse_transactions(
        self, source: _Report, report: ImportValidationReport
    ) -> list[RawTransaction]:
        parsed: list[RawTransaction] = []
        signatures: Counter[str] = Counter()
        for number, values in enumerate(source.rows, start=2):
            row = _Row(source.headers, values)
            if not any(_clean(v) for v in values):
                continue
            try:
                if not _clean(row.get("Datum", "Date")):
                    # DEGIRO sometimes wraps a long product name onto a
                    # continuation row without a date or transaction data.
                    report.rows_skipped += 1
                    continue
                occurred = _parse_datetime(
                    row.get("Datum", "Date"), row.get("Tijd", "Time")
                )
                product = _clean(row.get("Product"))
                isin = _clean(
                    row.get("ISIN", "Symbool/ISIN", "Symbol/ISIN")
                ).upper()
                order_id = _clean(row.get("Order ID", "Order Id", "OrderID"))
                venue = _clean(
                    row.get(
                        "Uitvoeringsplaats",
                        "Plaats van uitvoering",
                        "Venue",
                        "Execution venue",
                        "Beurs",
                        "Exchange",
                        "Referentiebeurs",
                        "Reference exchange",
                    )
                )
                quantity = _decimal(
                    row.get("Aantal", "Quantity"), required=True
                )
                price = _decimal(row.get("Koers", "Price"))
                fx_rate = _decimal(row.get("Wisselkoers", "Exchange rate"))
                fee = _decimal(
                    row.get(
                        "Transactiekosten",
                        "Transactiekosten en/of",
                        "Transactiekosten en/of kosten van derden",
                        "Transactiekosten en/of kosten van derden EUR",
                        "Transaction and/or third party fees",
                    )
                )
                autofx_fee = _decimal(
                    row.get(
                        "AutoFX Kosten",
                        "AutoFX fees",
                        "AutoFX costs",
                    )
                )
                total_fee = (
                    abs(fee or Decimal(0)) + abs(autofx_fee or Decimal(0))
                    if fee is not None or autofx_fee is not None
                    else None
                )
                fee_currency = _currency(
                    row.get("Valuta", "Currency", occurrence=2),
                    default="EUR",
                )
                currency = _currency(
                    row.get("Valuta", "Currency") or row.after("Koers", "Price")
                )
                eur_value = _decimal(
                    row.get("Waarde in EUR", "Value in EUR", "Totaal", "Total")
                )
                local_value = _decimal(
                    row.get("Lokale waarde", "Waarde", "Local value")
                )
                if eur_value is None:
                    eur_value = local_value if currency == "EUR" else None
                if (
                    eur_value is None
                    and price is not None
                    and quantity is not None
                ):
                    local_value = local_value or price * quantity
                    if fx_rate is not None and fx_rate != 0:
                        eur_value = local_value / fx_rate
                    else:
                        eur_value = local_value
                if eur_value is None:
                    message = "EUR-waarde ontbreekt"
                    raise ValueError(message)
                event = "purchase" if quantity and quantity > 0 else "sale"
                original_value = local_value or eur_value
                amount = (
                    -abs(original_value)
                    if event == "purchase"
                    else abs(original_value)
                )
                amount_in_base = (
                    -abs(eur_value) if event == "purchase" else abs(eur_value)
                )
                signature = _hash(
                    "transactions",
                    order_id,
                    occurred.isoformat(),
                    isin,
                    event,
                    quantity,
                    price,
                    currency,
                    eur_value,
                    fee,
                )
                duplicate = signatures[signature]
                signatures[signature] += 1
                external_id = _hash(signature, duplicate)
                parsed.append(
                    RawTransaction(
                        external_transaction_id=external_id,
                        external_account_id=self.external_account_id,
                        amount=amount,
                        currency_code=currency,
                        amount_in_base=amount_in_base,
                        base_currency_code="EUR",
                        fx_rate=fx_rate,
                        occurred_at=occurred,
                        booked_at=occurred,
                        description=f"{event.title()}: {product}",
                        transaction_type=event,
                        status="booked",
                        quantity=quantity,
                        unit_price=price,
                        fee_amount=total_fee,
                        fee_currency_code=fee_currency
                        if total_fee is not None
                        else None,
                        security_reference=SecurityReference(
                            external_id=isin or None,
                            isin=isin or None,
                            name=product or None,
                            venue=venue or None,
                            currency_code=currency,
                        ),
                        provider_fingerprint=signature,
                        provider_metadata={
                            "report_type": "transactions",
                            "order_id": order_id or None,
                            "price": str(price) if price is not None else None,
                            "instrument_currency": currency,
                            "local_value": str(local_value)
                            if local_value is not None
                            else None,
                            "transaction_fee": str(fee)
                            if fee is not None
                            else None,
                            "autofx_fee": str(autofx_fee)
                            if autofx_fee is not None
                            else None,
                            "total_fee": str(total_fee)
                            if total_fee is not None
                            else None,
                        },
                    )
                )
            except (ValueError, InvalidOperation) as exc:
                report.errors.append(
                    f"{source.source.name}, regel {number}: {exc}"
                )
        return parsed

    def _parse_statement(
        self, source: _Report, report: ImportValidationReport
    ) -> list[RawTransaction]:
        parsed: list[RawTransaction] = []
        signatures: Counter[str] = Counter()
        # DEGIRO records a foreign-currency dividend first, followed by a
        # separate ``Valuta Debitering`` and EUR ``Valuta Creditering``. The
        # technical rows are not economic transactions, but the debit carries
        # the exact historical FX rate needed for the EUR cash projection.
        pending_cash_fx: list[tuple[Decimal, Decimal]] = []
        latest_balance_at: datetime | None = None
        latest_balance: Decimal | None = None
        for number, values in enumerate(source.rows, start=2):
            row = _Row(source.headers, values)
            try:
                description = _clean(row.get("Omschrijving", "Description"))
                if not description and not _clean(row.get("Datum", "Date")):
                    continue
                lowered = description.casefold()
                if self._is_technical_statement_row(lowered):
                    if "valuta debitering" in lowered:
                        fx = _decimal(
                            row.get(
                                "FX mutatie",
                                "Wisselkoers",
                                "FX",
                                "Exchange rate",
                            )
                        )
                        mutation_raw = row.get("Mutatie", "Change", "Amount")
                        mutation_currency = _currency(mutation_raw, default="")
                        if mutation_currency:
                            mutation_raw = row.after(
                                "Mutatie", "Change", "Amount"
                            )
                        mutation = _decimal(mutation_raw)
                        if (
                            mutation is not None
                            and mutation_currency == "USD"
                            and fx not in (None, 0)
                        ):
                            pending_cash_fx.append((abs(mutation), fx))
                    report.rows_skipped += 1
                    continue
                occurred = _parse_datetime(
                    row.get("Datum", "Date"), row.get("Tijd", "Time")
                )
                balance_raw = row.get("Saldo", "Balance")
                balance_currency = _currency(balance_raw, default="")
                if balance_currency:
                    balance_raw = row.after("Saldo", "Balance")
                balance = _decimal(balance_raw)
                if balance is not None and (
                    latest_balance_at is None or occurred >= latest_balance_at
                ):
                    latest_balance_at = occurred
                    latest_balance = balance
                order_id = _clean(row.get("Order ID", "Order Id", "OrderID"))
                if order_id in self._trade_order_ids and self._is_trade_fee(
                    lowered
                ):
                    report.rows_skipped += 1
                    continue
                isin = _clean(row.get("ISIN")).upper()
                product = _clean(row.get("Product"))
                mutation_raw = row.get("Mutatie", "Change", "Amount")
                # Current DEGIRO XLSX files put the currency immediately
                # before the numeric mutation (with an unlabeled spacer
                # column). Older exports put the numeric value directly in
                # the named column, so support both layouts.
                mutation_currency = _currency(mutation_raw, default="")
                if mutation_currency:
                    mutation_raw = row.after("Mutatie", "Change", "Amount")
                if mutation_raw in (None, ""):
                    # Balance-only rows are common in the Account export
                    # (for example the cash-sweep header rows). They update
                    # the account balance but are not transactions.
                    report.rows_skipped += 1
                    continue
                amount = _decimal(mutation_raw, required=True)
                mutation_after = row.after("Mutatie", "Change", "Amount")
                currency = _currency(
                    row.get("Valuta", "Currency", occurrence=1)
                    or _currency(mutation_after, default="")
                    or mutation_currency
                    or row.get("Valuta", "Currency")
                )
                fx = _decimal(
                    row.get(
                        "FX mutatie",
                        "Wisselkoers",
                        "FX",
                        "Exchange rate",
                    )
                )
                transaction_type = self._statement_type(
                    lowered, amount or Decimal(0)
                )
                projected_fx = None
                if currency != "EUR" and transaction_type in {
                    "dividend",
                    "interest",
                    "deposit",
                    "withdrawal",
                    "fee",
                    "tax",
                }:
                    for index, (remaining, rate) in enumerate(pending_cash_fx):
                        # One DEGIRO FX debit can settle several dividends
                        # together (for example USD 124.34 covering two
                        # dividend rows). Allocate the same broker rate to
                        # each component until the debit is consumed.
                        if remaining < abs(amount):
                            continue
                        projected_fx = rate
                        leftover = remaining - abs(amount)
                        if leftover:
                            pending_cash_fx[index] = (leftover, rate)
                        else:
                            pending_cash_fx.pop(index)
                        break
                    if projected_fx is not None:
                        fx = projected_fx
                signature = _hash(
                    "account_statement",
                    order_id,
                    occurred.isoformat(),
                    isin,
                    transaction_type,
                    amount,
                    currency,
                    description,
                )
                duplicate = signatures[signature]
                signatures[signature] += 1
                parsed.append(
                    RawTransaction(
                        external_transaction_id=_hash(signature, duplicate),
                        external_account_id=self.external_account_id,
                        amount=amount or Decimal(0),
                        currency_code=currency,
                        occurred_at=occurred,
                        booked_at=occurred,
                        description=description,
                        transaction_type=transaction_type,
                        status="booked",
                        amount_in_base=(
                            amount / fx
                            if amount is not None and fx not in (None, 0)
                            else None
                        ),
                        base_currency_code="EUR"
                        if fx not in (None, 0)
                        else None,
                        security_reference=SecurityReference(
                            external_id=isin or None,
                            isin=isin or None,
                            name=product or None,
                            currency_code=currency,
                        )
                        if isin
                        else None,
                        fx_rate=fx,
                        provider_fingerprint=signature,
                        provider_metadata={
                            "report_type": "account_statement",
                            "order_id": order_id or None,
                            "source_currency": currency,
                            "fx_rate": str(fx) if fx is not None else None,
                            "fx_projection_source": (
                                "paired_valuta_debitering"
                                if projected_fx is not None
                                else None
                            ),
                        },
                    )
                )
            except (ValueError, InvalidOperation) as exc:
                report.errors.append(
                    f"{source.source.name}, regel {number}: {exc}"
                )
        # An Account/statement export has no positions, but it does contain
        # the latest cash balance. Preserve it as the account balance when no
        # portfolio export was supplied.
        if self._portfolio_total is None and latest_balance is not None:
            self._portfolio_total = latest_balance
            self._cash_total = latest_balance
        return parsed

    @staticmethod
    def _is_technical_statement_row(description: str) -> bool:
        technical = (
            "cash sweep",
            "geldmarktfonds",
            "flatex cash",
            "fx debit",
            "fx credit",
            "valuta debitering",
            "valuta creditering",
        )
        trade_mirror = ("koop ", "aankoop ", "buy ", "verkoop ", "sell ")
        return any(term in description for term in technical + trade_mirror)

    @staticmethod
    def _is_trade_fee(description: str) -> bool:
        return "transactiekosten" in description or (
            "transaction" in description and "fee" in description
        )

    @staticmethod
    def _statement_type(description: str, amount: Decimal) -> str:
        mappings = (
            (("dividendbelasting", "dividend tax", "withholding tax"), "tax"),
            (("dividend",), "dividend"),
            (("storting", "deposit"), "deposit"),
            (("opname", "withdrawal"), "withdrawal"),
            (("rente", "interest"), "interest"),
            (
                (
                    "valutatransactie",
                    "currency exchange",
                    "fx settlement",
                    "fx conversie",
                ),
                "transfer",
            ),
            (
                ("kosten", "fee", "aansluit", "platform", "corporate action"),
                "fee",
            ),
        )
        for terms, kind in mappings:
            if any(term in description for term in terms):
                return kind
        return "deposit" if amount > 0 else "withdrawal"

    def _parse_portfolio(
        self, source: _Report, report: ImportValidationReport
    ) -> list[RawHolding]:
        holdings: list[RawHolding] = []
        snapshot = datetime.fromtimestamp(source.source.stat().st_mtime, tz=UTC)
        configured = self.config.options.get("snapshot_at")
        if configured:
            snapshot = _parse_datetime(configured)
        total = Decimal(0)
        cash_total = Decimal(0)
        for number, values in enumerate(source.rows, start=2):
            row = _Row(source.headers, values)
            try:
                product = _clean(row.get("Product"))
                isin = _clean(
                    row.get("ISIN", "Symbool/ISIN", "Symbol/ISIN")
                ).upper()
                quantity = _decimal(row.get("Aantal", "Quantity"))
                price = _decimal(
                    row.get("Slotkoers", "Closing price", "Koers", "Price")
                )
                currency = _currency(
                    row.get("Valuta", "Currency")
                    or row.after("Slotkoers", "Closing price", "Koers", "Price")
                )
                market_value = _decimal(
                    row.get("Waarde in EUR", "Value in EUR")
                )
                # Some current DEGIRO Portfolio CSV exports label the currency
                # column as "Lokale waarde", leave the numeric local-value
                # column blank, and put "Waarde in EUR" after it. In that
                # layout the normal alias returns "EUR" or "USD".
                local_value_raw = row.get(
                    "Lokale waarde", "Local value", "Waarde", "Value"
                )
                if _currency(local_value_raw, default=""):
                    local_value_raw = row.after(
                        "Slotkoers",
                        "Closing price",
                        "Koers",
                        "Price",
                        offset=2,
                    )
                local_value = _decimal(local_value_raw)
                market_value = (
                    market_value if market_value is not None else local_value
                )
                cost_basis = _decimal(
                    row.get(
                        "GAK",
                        "Gemiddelde aankoopkoers",
                        "Average price",
                        "Average purchase price",
                        "Cost basis",
                    )
                )
                if not product and not isin and quantity is None:
                    continue
                cash_keys = {"eur", "cash", "geldrekening", "cashaccount"}
                product_key = _key(product)
                is_cash = _key(isin) in cash_keys or (
                    not isin
                    and (
                        product_key in cash_keys
                        or "cash" in product_key
                        or "geldrekening" in product_key
                    )
                )
                if is_cash:
                    cash = (
                        market_value
                        if market_value is not None
                        else quantity or Decimal(0)
                    )
                    cash_total += cash
                    total += cash
                    report.rows_skipped += 1
                    continue
                if not isin:
                    message = "ISIN ontbreekt voor effect"
                    raise ValueError(message)
                if quantity is None:
                    message = "aantal ontbreekt"
                    raise ValueError(message)
                total += market_value or Decimal(0)
                holdings.append(
                    RawHolding(
                        external_account_id=self.external_account_id,
                        observed_at=snapshot,
                        quantity=quantity,
                        security_reference=SecurityReference(
                            external_id=isin,
                            isin=isin,
                            name=product or None,
                            currency_code=currency,
                        ),
                        cost_basis=(cost_basis * quantity)
                        if cost_basis is not None
                        else None,
                        cost_basis_currency=currency,
                        market_value=market_value,
                        currency_code="EUR"
                        if market_value is not None
                        else currency,
                        price=price,
                        price_currency=currency,
                        provider_metadata={
                            "report_type": "portfolio",
                            "local_value": str(local_value)
                            if local_value is not None
                            else None,
                            "average_price": str(cost_basis)
                            if cost_basis is not None
                            else None,
                        },
                    )
                )
            except (ValueError, InvalidOperation) as exc:
                report.errors.append(
                    f"{source.source.name}, regel {number}: {exc}"
                )
        if self._snapshot_at is None or snapshot >= self._snapshot_at:
            self._snapshot_at = snapshot
            self._portfolio_total = total
            self._cash_total = cash_total
        return holdings

    @staticmethod
    def _deduplicate(items: list[RawTransaction]) -> list[RawTransaction]:
        unique: dict[str, RawTransaction] = {}
        for item in items:
            unique.setdefault(item.external_transaction_id, item)
        return sorted(
            unique.values(),
            key=lambda item: (item.occurred_at, item.external_transaction_id),
        )
