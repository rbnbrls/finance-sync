"""Secure staging, preview and execution for DEGIRO export imports."""

from __future__ import annotations

import contextlib
import csv
import hashlib
import io
import json
import os
import re
import shutil
import zipfile
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING, Any, cast
from uuid import uuid4

from sqlalchemy import func, select

from finance_sync.connectors.degiro_pension import DegiroPensionConnector
from finance_sync.connectors.models import ConnectorConfig
from finance_sync.connectors.registry import ConnectorRegistry
from finance_sync.models.account import Account
from finance_sync.models.transaction import Transaction
from finance_sync.services.auth import encrypt_credential
from finance_sync.sync.orchestrator import SyncOrchestrator

if TYPE_CHECKING:
    from fastapi import UploadFile
    from sqlalchemy.ext.asyncio import AsyncSession

    from finance_sync.config.settings import Settings
    from finance_sync.container import Container
    from finance_sync.models.credential import Credential
    from finance_sync.models.import_run import ImportRun

_SUPPORTED_SUFFIXES = {".csv", ".xlsx", ".xls"}
_FORMULA_PREFIXES = ("=", "+", "@")
_EXPECTED_REPORTS = {"transactions", "account_statement", "portfolio"}
_SAFE_FILE_RE = re.compile(r"[^A-Za-z0-9._-]+")


class ImportValidationError(ValueError):
    """A safe validation error suitable for returning to an administrator."""


def connector_options(credential: Credential) -> dict[str, Any]:
    """Decode the non-secret options stored in the legacy description field."""
    try:
        value = json.loads(credential.description or "{}")
    except (json.JSONDecodeError, TypeError):
        return {}
    if not isinstance(value, dict):
        return {}
    result = dict(cast("dict[str, Any]", value))
    result.pop("_label", None)
    return result


def _safe_name(filename: str | None, index: int) -> tuple[str, str]:
    supplied = filename or f"export-{index}.csv"
    if Path(supplied).name != supplied or "\x00" in supplied:
        message = "Een bestandsnaam bevat een ongeldig pad."
        raise ImportValidationError(message)
    suffix = Path(supplied).suffix.lower()
    if suffix not in _SUPPORTED_SUFFIXES:
        message = (
            "Alleen DEGIRO-exports in CSV-, XLSX- of XLS-formaat "
            "zijn toegestaan."
        )
        raise ImportValidationError(message)
    display = (
        _SAFE_FILE_RE.sub("_", supplied)[:160] or f"export-{index}{suffix}"
    )
    return display, suffix


def _check_xlsx_archive(path: Path, max_bytes: int) -> None:
    try:
        with zipfile.ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > 200:
                message = "Het Excel-bestand bevat onverwacht veel onderdelen."
                raise ImportValidationError(message)
            expanded = sum(entry.file_size for entry in entries)
            compressed = sum(max(entry.compress_size, 1) for entry in entries)
            if expanded > max_bytes * 10 or expanded / compressed > 100:
                message = "Het Excel-bestand is ongewoon sterk gecomprimeerd."
                raise ImportValidationError(message)
            if any(".." in Path(entry.filename).parts for entry in entries):
                message = "Het Excel-bestand bevat onveilige interne paden."
                raise ImportValidationError(message)
    except zipfile.BadZipFile as exc:
        message = "Het XLSX-bestand is beschadigd of geen geldig Excel-bestand."
        raise ImportValidationError(message) from exc


def _formula_like(value: object) -> bool:
    if not isinstance(value, str):
        return False
    stripped = value.lstrip()
    return stripped.startswith(_FORMULA_PREFIXES) or (
        stripped.startswith("-")
        and not re.fullmatch(r"-\d+(?:[.,]\d+)?", stripped)
    )


def _check_formula_injection(path: Path) -> None:
    suffix = path.suffix.lower()
    rows: Any
    if suffix == ".csv":
        data = path.read_bytes()
        text = None
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                text = data.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if text is None:
            message = "De CSV-tekstcodering wordt niet ondersteund."
            raise ImportValidationError(message)
        sample = text[:8192]
        delimiter = ";" if sample.count(";") > sample.count(",") else ","
        rows = csv.reader(io.StringIO(text), delimiter=delimiter)
    elif suffix == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook.active
            rows = [] if sheet is None else sheet.iter_rows(values_only=True)
            if any(_formula_like(cell) for row in rows for cell in row):
                message = (
                    "Het bestand bevat formules; upload een ongewijzigde "
                    "DEGIRO-export."
                )
                raise ImportValidationError(message)
            return
        finally:
            workbook.close()
    else:
        import xlrd

        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            sheet = book.sheet_by_index(0)
            rows = (sheet.row_values(index) for index in range(sheet.nrows))
            if any(_formula_like(cell) for row in rows for cell in row):
                message = (
                    "Het bestand bevat formules; upload een ongewijzigde "
                    "DEGIRO-export."
                )
                raise ImportValidationError(message)
            return
        finally:
            book.release_resources()
    if any(_formula_like(cell) for row in rows for cell in row):
        message = "Het bestand bevat formule-achtige inhoud en is geweigerd."
        raise ImportValidationError(message)


def _tenant_stage(settings: Settings, tenant_id: str, run_id: str) -> Path:
    root = settings.degiro_import_staging_directory.resolve()
    path = root / hashlib.sha256(tenant_id.encode()).hexdigest()[:24] / run_id
    path.mkdir(parents=True, mode=0o700, exist_ok=False)
    os.chmod(path, 0o700)
    return path


async def stage_uploads(
    files: list[UploadFile],
    *,
    settings: Settings,
    tenant_id: str,
    run_id: str,
) -> tuple[list[Path], list[str], list[str]]:
    """Stream uploads into a private tenant/run directory with hard limits."""
    if not files or len(files) > settings.degiro_import_max_files:
        message = (
            f"Selecteer 1 tot {settings.degiro_import_max_files} bestanden."
        )
        raise ImportValidationError(message)
    directory = _tenant_stage(settings, tenant_id, run_id)
    paths: list[Path] = []
    display_names: list[str] = []
    hashes: list[str] = []
    batch_size = 0
    try:
        for index, upload in enumerate(files, start=1):
            display, suffix = _safe_name(upload.filename, index)
            destination = directory / f"{index:02d}-{uuid4().hex}{suffix}"
            digest = hashlib.sha256()
            size = 0
            with destination.open("xb") as output:
                while chunk := await upload.read(64 * 1024):
                    size += len(chunk)
                    batch_size += len(chunk)
                    if size > settings.degiro_import_max_file_bytes:
                        message = (
                            f"{display} is groter dan de ingestelde "
                            "uploadlimiet."
                        )
                        raise ImportValidationError(message)
                    if batch_size > settings.degiro_import_max_batch_bytes:
                        message = (
                            "De totale uploadbatch overschrijdt de limiet."
                        )
                        raise ImportValidationError(message)
                    digest.update(chunk)
                    output.write(chunk)
            if size == 0:
                message = f"{display} is leeg."
                raise ImportValidationError(message)
            if suffix == ".xlsx":
                _check_xlsx_archive(
                    destination, settings.degiro_import_max_file_bytes
                )
            if suffix == ".xls" and destination.read_bytes()[
                :8
            ] != bytes.fromhex("D0CF11E0A1B11AE1"):
                message = f"{display} is geen geldig XLS-bestand."
                raise ImportValidationError(message)
            _check_formula_injection(destination)
            paths.append(destination)
            display_names.append(display)
            hashes.append(digest.hexdigest())
        return paths, display_names, hashes
    except Exception:
        shutil.rmtree(directory, ignore_errors=True)
        raise


def batch_hash(hashes: list[str]) -> str:
    return hashlib.sha256("\n".join(sorted(hashes)).encode()).hexdigest()


def validate_local_files(paths: list[Path], settings: Settings) -> None:
    """Apply upload-equivalent safety limits to claimed watchfolder files."""
    if not paths or len(paths) > settings.degiro_import_max_files:
        message = "De watchfolderbatch bevat een ongeldig aantal bestanden."
        raise ImportValidationError(message)
    total_size = 0
    for path in paths:
        suffix = path.suffix.lower()
        if suffix not in _SUPPORTED_SUFFIXES or not path.is_file():
            message = "De watchfolderbatch bevat een niet-ondersteund bestand."
            raise ImportValidationError(message)
        file_size = path.stat().st_size
        total_size += file_size
        if file_size > settings.degiro_import_max_file_bytes:
            message = "Een watchfolderbestand overschrijdt de bestandlimiet."
            raise ImportValidationError(message)
        if total_size > settings.degiro_import_max_batch_bytes:
            message = "De totale watchfolderbatch overschrijdt de limiet."
            raise ImportValidationError(message)
        if suffix == ".xlsx":
            _check_xlsx_archive(path, settings.degiro_import_max_file_bytes)
        if suffix == ".xls" and path.read_bytes()[:8] != bytes.fromhex(
            "D0CF11E0A1B11AE1"
        ):
            message = "De watchfolderbatch bevat een ongeldig XLS-bestand."
            raise ImportValidationError(message)
        _check_formula_injection(path)


async def build_preview(
    paths: list[Path],
    *,
    options: dict[str, Any],
    settings: Settings,
    session: AsyncSession,
    tenant_id: str,
) -> dict[str, Any]:
    connector = DegiroPensionConnector(
        ConnectorConfig(
            provider_type="degiro_pension",
            options={**options, "export_paths": [str(path) for path in paths]},
        )
    )
    await connector.authenticate()
    report = connector.validation_report
    if report.rows_read > settings.degiro_import_max_rows:
        message = (
            f"De export bevat {report.rows_read} regels; de limiet is "
            f"{settings.degiro_import_max_rows}."
        )
        raise ImportValidationError(message)
    transactions = await connector.fetch_transactions(
        datetime.min.replace(tzinfo=UTC)
    )
    holdings = await connector.fetch_holdings()
    account = (await connector.fetch_accounts())[0]
    dates = [item.occurred_at for item in transactions]
    tx_ids = [item.external_transaction_id for item in transactions]
    duplicate_count = 0
    if tx_ids:
        duplicate_count = int(
            (
                await session.execute(
                    select(func.count(Transaction.id)).where(
                        Transaction.tenant_id == tenant_id,
                        Transaction.provider_key == "degiro_pension",
                        Transaction.external_transaction_id.in_(tx_ids),
                    )
                )
            ).scalar_one()
        )
    unresolved = sorted(
        {
            reference.isin or reference.name or "onbekend instrument"
            for reference in [
                *(item.security_reference for item in transactions),
                *(item.security_reference for item in holdings),
            ]
            if reference is not None and not reference.isin
        }
    )
    report_types = sorted(report.report_types)
    return {
        "reports": connector.report_summaries,
        "report_types": report_types,
        "missing_report_types": sorted(_EXPECTED_REPORTS - set(report_types)),
        "account_label": account.name,
        "external_account_id": account.external_account_id,
        "period_start": min(dates).isoformat() if dates else None,
        "period_end": max(dates).isoformat() if dates else None,
        "rows": report.rows_read,
        "transactions": len(transactions),
        "holdings": len(holdings),
        "skipped": report.rows_skipped,
        "possible_duplicates": duplicate_count,
        "unknown_transaction_types": sorted(
            {
                item.transaction_type
                for item in transactions
                if item.transaction_type is not None
                and item.transaction_type
                not in {
                    "purchase",
                    "sale",
                    "deposit",
                    "withdrawal",
                    "dividend",
                    "tax",
                    "interest",
                    "fee",
                    "transfer",
                    "other",
                }
            }
        ),
        "unresolved_securities": unresolved,
        "warnings": list(report.warnings),
    }


def stage_paths(
    settings: Settings, tenant_id: str, run: ImportRun
) -> list[Path]:
    base = (
        settings.degiro_import_staging_directory.resolve()
        / hashlib.sha256(tenant_id.encode()).hexdigest()[:24]
        / str(run.id)
    )
    return [base / name for name in run.storage_names]


def verify_staged(run: ImportRun, paths: list[Path]) -> None:
    if len(paths) != len(run.content_hashes):
        message = "De gevalideerde upload is niet meer compleet."
        raise ImportValidationError(message)
    actual: list[str] = []
    for path in paths:
        if not path.is_file():
            message = "De gevalideerde upload is verlopen of verwijderd."
            raise ImportValidationError(message)
        actual.append(hashlib.sha256(path.read_bytes()).hexdigest())
    if actual != run.content_hashes:
        message = (
            "De upload is na de preview gewijzigd en kan niet worden bevestigd."
        )
        raise ImportValidationError(message)


async def execute_run(
    run: ImportRun,
    *,
    paths: list[Path],
    options: dict[str, Any],
    container: Container,
    session: AsyncSession,
    retain: bool = False,
    cleanup: bool = True,
) -> ImportRun:
    verify_staged(run, paths)
    run.status = "processing"
    await session.flush()
    config = ConnectorConfig(
        provider_type="degiro_pension",
        options={**options, "export_paths": [str(path) for path in paths]},
    )
    try:
        orchestrator = SyncOrchestrator(
            session_factory=container.session_factory,
            registry=ConnectorRegistry(),
            tenant_id=str(run.tenant_id),
            settings=container.settings,
        )
        result = await orchestrator.run_sync(
            provider_type="degiro_pension",
            config=config,
            since=datetime.min.replace(tzinfo=UTC),
        )
        if result.status.value == "failed":
            message = "De import kon niet atomair worden verwerkt."
            raise ImportValidationError(message)
        run.status = "completed"
        processed = (
            result.accounts_synced
            + result.transactions_synced
            + result.holdings_synced
        )
        run.updated_count = min(
            processed, int(run.preview.get("possible_duplicates", 0))
        )
        run.created_count = processed - run.updated_count
        run.skipped_count = int(run.preview.get("skipped", 0))
        run.completed_at = datetime.now(UTC)
        account = (
            await session.execute(
                select(Account).where(
                    Account.tenant_id == run.tenant_id,
                    Account.provider_key == "degiro_pension",
                    Account.external_account_id
                    == run.preview.get("external_account_id"),
                )
            )
        ).scalar_one_or_none()
        if account is not None:
            run.account_id = account.id
        run.audit_events = [
            *run.audit_events,
            {"action": "confirmed", "at": datetime.now(UTC).isoformat()},
        ]
        if retain:
            _retain_encrypted(run, paths, container.settings)
            run.retained = True
        return run
    except Exception as exc:
        run.status = "failed"
        run.rejected_count = run.rows_total
        run.completed_at = datetime.now(UTC)
        run.error_details = [_sanitized_error(exc)]
        raise
    finally:
        if cleanup and not retain:
            for path in paths:
                with contextlib.suppress(OSError):
                    path.unlink()
            if paths:
                shutil.rmtree(paths[0].parent, ignore_errors=True)


def _retain_encrypted(
    run: ImportRun, paths: list[Path], settings: Settings
) -> None:
    retained = (
        settings.degiro_import_staging_directory / "retained" / str(run.id)
    )
    retained.mkdir(parents=True, mode=0o700)
    for index, path in enumerate(paths):
        plaintext = path.read_bytes().hex()
        ciphertext, nonce = encrypt_credential(plaintext, settings)
        (retained / f"{index:02d}.enc").write_bytes(nonce + ciphertext)
    shutil.rmtree(paths[0].parent, ignore_errors=True)


def _sanitized_error(exc: Exception) -> str:
    if isinstance(exc, ImportValidationError):
        return str(exc)[:500]
    return "De import is mislukt. Controleer de validatiewaarschuwingen."


def cleanup_expired_previews(settings: Settings) -> int:
    root = settings.degiro_import_staging_directory
    if not root.exists():
        return 0
    cutoff = datetime.now(UTC) - timedelta(
        minutes=settings.degiro_import_preview_ttl_minutes
    )
    removed = 0
    for tenant_dir in root.iterdir():
        if not tenant_dir.is_dir() or tenant_dir.name == "retained":
            continue
        for run_dir in tenant_dir.iterdir():
            modified = datetime.fromtimestamp(run_dir.stat().st_mtime, tz=UTC)
            if run_dir.is_dir() and modified < cutoff:
                shutil.rmtree(run_dir, ignore_errors=True)
                removed += 1
    return removed
