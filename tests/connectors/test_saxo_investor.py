"""Tests for SaxoInvestor XLSX position snapshots."""

from __future__ import annotations

from datetime import UTC, datetime
from decimal import Decimal
from typing import TYPE_CHECKING

import pytest
from openpyxl import Workbook

from finance_sync.connectors.exceptions import PermanentError
from finance_sync.connectors.models import ConnectorConfig
from finance_sync.connectors.saxo_investor import SaxoInvestorConnector

if TYPE_CHECKING:
    from pathlib import Path

HEADERS = [
    "Instrument",
    "Valuta",
    "Aantal",
    "Actuele koers",
    "Huidige waarde (EUR)",
    "Kostprijs",
    "Symbool",
    "ISIN",
    "Soort belegging",
]

TRANSACTION_HEADERS = [
    "Transactiedatum",
    "Valutadatum",
    "Rekening-ID",
    "Transactie-ID",
    "Bk Record Id",
    "Booking Id",
    "Transactietype",
    "Acties",
    "Boekingsbedrag",
    "Valuta",
    "Totale kosten",
    "Instrument",
    "Instrumentsymbool",
    "Instrument ISIN",
    "Instrumentvaluta",
    "Type",
]


def _write_export(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(HEADERS)
    sheet.append(
        [
            "EUR (1)",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    sheet.append(
        [
            "Example Equity",
            "EUR",
            2,
            100.5,
            201,
            90.25,
            "EXM:xams",
            "NL0000000001",
            "Aandeel",
        ]
    )
    sheet.append(
        [
            "Example ETF",
            "USD",
            3,
            20,
            55,
            18,
            "ETF:xnas",
            "IE0000000002",
            "ETF's",
        ]
    )
    workbook.save(path)


def _write_transactions(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Transacties"
    sheet.append(TRANSACTION_HEADERS)
    sheet.append(
        [
            datetime(2026, 4, 22),
            datetime(2026, 4, 24),
            "15996986",
            123,
            123,
            456,
            "Transactie",
            "Koop 1000 @ 1.91 USD",
            -1633.37,
            "EUR",
            -4.92,
            "Ready Capital Corp",
            "RC:xnys",
            "US75574V1016",
            "USD",
            "Aandeel",
        ]
    )
    sheet.append(
        [
            datetime(2026, 4, 17),
            datetime(2026, 4, 21),
            "15996986",
            124,
            124,
            457,
            "Transactie",
            "Verkoop -107 @ 18.10 USD",
            1632.43,
            "EUR",
            -4.96,
            "Petrobras",
            "PBR:xnys",
            "US71654V1017",
            "USD",
            "Aandeel",
        ]
    )
    workbook.save(path)


@pytest.mark.asyncio
async def test_imports_saxo_positions_as_holdings(tmp_path: Path) -> None:
    path = tmp_path / "Posities_23-aug-2026_12_56_37.xlsx"
    _write_export(path)
    connector = SaxoInvestorConnector(
        ConnectorConfig(
            provider_type="saxo_investor",
            options={"export_path": str(path), "account_key": "my-saxo"},
        )
    )

    await connector.authenticate()
    account = (await connector.fetch_accounts())[0]
    holdings = await connector.fetch_holdings()

    assert connector.name == "saxo_investor"
    assert connector.supported_resources == {
        "accounts",
        "transactions",
        "holdings",
    }
    assert account.external_account_id == "saxo-investor-my-saxo"
    assert account.current_balance == Decimal(256)
    assert len(holdings) == 2
    assert holdings[0].observed_at == datetime(2026, 8, 23, tzinfo=UTC)
    assert holdings[0].security_reference.isin == "NL0000000001"
    assert holdings[0].security_reference.security_type == "equity"
    assert holdings[1].price_currency == "USD"
    assert holdings[1].currency_code == "EUR"
    assert holdings[1].market_value == Decimal(55)
    assert (
        await connector.fetch_transactions(datetime.min.replace(tzinfo=UTC))
        == []
    )


@pytest.mark.asyncio
async def test_snapshot_at_overrides_filename_and_account_filter(
    tmp_path: Path,
) -> None:
    path = tmp_path / "positions.xlsx"
    _write_export(path)
    connector = SaxoInvestorConnector(
        ConnectorConfig(
            provider_type="saxo_investor",
            options={
                "export_path": str(path),
                "snapshot_at": "2026-08-24",
            },
        )
    )
    await connector.authenticate()
    account_id = (await connector.fetch_accounts())[0].external_account_id
    assert (await connector.fetch_holdings(account_id="other")) == []
    assert (await connector.fetch_holdings(account_id=account_id))[
        0
    ].observed_at == datetime(2026, 8, 24, tzinfo=UTC)


@pytest.mark.asyncio
async def test_imports_saxo_transactions_and_combines_both_exports(
    tmp_path: Path,
) -> None:
    positions = tmp_path / "Posities_23-aug-2026.xlsx"
    transactions = tmp_path / "Transactions_2026-01-01_2026-08-23.xlsx"
    _write_export(positions)
    _write_transactions(transactions)
    connector = SaxoInvestorConnector(
        ConnectorConfig(
            provider_type="saxo_investor",
            options={"export_paths": [str(positions), str(transactions)]},
        )
    )

    await connector.authenticate()
    imported = await connector.fetch_transactions(
        datetime.min.replace(tzinfo=UTC)
    )

    assert len(imported) == 2
    assert imported[0].transaction_type == "purchase"
    assert imported[0].security_reference.isin == "US75574V1016"
    assert imported[0].fee_amount == Decimal("4.92")
    assert len(await connector.fetch_holdings()) == 2


@pytest.mark.asyncio
async def test_accepts_zero_booking_amount(tmp_path: Path) -> None:
    path = tmp_path / "Transactions_zero.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Transacties"
    sheet.append(TRANSACTION_HEADERS)
    sheet.append(
        [
            datetime(2026, 5, 29),
            datetime(2026, 5, 28),
            "15996986",
            None,
            2335654579,
            None,
            "Corporate action",
            "Dividend",
            0,
            "EUR",
            -0.48,
            "Petroleo Brasileiro SA Petrobras - Pref ADR",
            "PBRa:xnys",
            "US71654V1017",
            "USD",
            "Stock",
        ]
    )
    workbook.save(path)

    connector = SaxoInvestorConnector(
        ConnectorConfig(
            provider_type="saxo_investor", options={"export_path": str(path)}
        )
    )

    await connector.authenticate()
    imported = await connector.fetch_transactions(
        datetime.min.replace(tzinfo=UTC)
    )

    assert len(imported) == 1
    assert imported[0].amount == 0


@pytest.mark.asyncio
async def test_skips_transaction_rows_with_missing_booking_amount(
    tmp_path: Path,
) -> None:
    """A summary/separator row without a booking amount must not abort the
    import (GlitchTip #6 / GitHub #463 regression)."""
    path = tmp_path / "Transactions_missing_amount.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Transacties"
    sheet.append(TRANSACTION_HEADERS)
    # Valid transaction row.
    sheet.append(
        [
            datetime(2026, 4, 22),
            datetime(2026, 4, 24),
            "15996986",
            123,
            123,
            456,
            "Transactie",
            "Koop 1000 @ 1.91 USD",
            -1633.37,
            "EUR",
            -4.92,
            "Ready Capital Corp",
            "RC:xnys",
            "US75574V1016",
            "USD",
            "Aandeel",
        ]
    )
    # Summary/separator row: no booking amount (empty cell).
    sheet.append(
        [
            datetime(2026, 4, 30),
            None,
            None,
            None,
            None,
            None,
            None,
            "Totaal",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    # Summary/separator row rendered with a dash instead of a number.
    sheet.append(
        [
            datetime(2026, 4, 30),
            None,
            None,
            None,
            None,
            None,
            None,
            "Subtotaal",
            "-",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    workbook.save(path)

    connector = SaxoInvestorConnector(
        ConnectorConfig(
            provider_type="saxo_investor", options={"export_path": str(path)}
        )
    )

    await connector.authenticate()
    imported = await connector.fetch_transactions(
        datetime.min.replace(tzinfo=UTC)
    )
    account = (await connector.fetch_accounts())[0]

    assert len(imported) == 1
    assert imported[0].amount == Decimal("-1633.37")
    assert imported[0].security_reference is not None
    assert imported[0].security_reference.isin == "US75574V1016"
    assert account.provider_metadata is not None
    assert account.provider_metadata.get("skipped_transaction_rows") == 2
    assert account.provider_metadata.get("transactions_count") == 1


@pytest.mark.asyncio
async def test_summary_only_transactions_export_does_not_abort_positions(
    tmp_path: Path,
) -> None:
    """A transactions sheet with only summary/separator rows (empty trading
    period) must not abort the import of a valid positions export
    (GlitchTip #6 / GitHub #463 regression).

    Saxo exports still emit the transaction header followed by subtotal,
    account-total and blank separator rows when the period has no trades.
    Pre-fix, the first such row raised ``Boekingsbedrag op regel N
    ontbreekt`` inside ``_parse_transactions`` and aborted the entire
    ``authenticate()`` — taking the positions of every account in the same
    run down with it. Post-fix the rows are skipped, zero transactions are
    imported, the positions still load, and the skipped count is surfaced
    in the account's ``provider_metadata``.
    """
    positions = tmp_path / "Posities_23-aug-2026.xlsx"
    transactions = tmp_path / "Transactions_2026-01-01_2026-08-23.xlsx"
    _write_export(positions)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Transacties"
    sheet.append(TRANSACTION_HEADERS)
    # Summary row with a date and currency but no booking amount — the exact
    # shape of production row 96 that failed with 'Boekingsbedrag op regel
    # 96 ontbreekt'.
    sheet.append(
        [
            datetime(2026, 8, 21),
            None,
            None,
            None,
            None,
            None,
            None,
            "Totaal",
            None,
            "EUR",
        ]
    )
    # Summary row rendered with a dash placeholder in the amount column.
    sheet.append(
        [
            datetime(2026, 8, 22),
            None,
            None,
            None,
            None,
            None,
            None,
            "Subtotaal",
            "-",
            "EUR",
        ]
    )
    workbook.save(transactions)

    connector = SaxoInvestorConnector(
        ConnectorConfig(
            provider_type="saxo_investor",
            options={"export_paths": [str(positions), str(transactions)]},
        )
    )

    # Pre-fix this raised PermanentError and lost the positions import too.
    await connector.authenticate()
    imported = await connector.fetch_transactions(
        datetime.min.replace(tzinfo=UTC)
    )
    holdings = await connector.fetch_holdings()
    account = (await connector.fetch_accounts())[0]

    assert imported == []
    assert len(holdings) == 2
    assert account.provider_metadata.get("transactions_count") == 0
    assert account.provider_metadata.get("skipped_transaction_rows") == 2


@pytest.mark.asyncio
async def test_rejects_non_saxo_layout(tmp_path: Path) -> None:
    path = tmp_path / "wrong.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.append(["Date", "Amount"])
    workbook.save(path)
    connector = SaxoInvestorConnector(
        ConnectorConfig(
            provider_type="saxo_investor", options={"export_path": str(path)}
        )
    )
    with pytest.raises(PermanentError, match="vereiste kolommen"):
        await connector.authenticate()
